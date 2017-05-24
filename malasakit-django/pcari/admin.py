from django.contrib import admin

from django.http import HttpResponse
from django.core import serializers
from django.contrib.contenttypes.models import ContentType
from django.http import HttpResponseRedirect

import csv
from django.utils.encoding import smart_str
import openpyxl
from openpyxl.utils import get_column_letter

# Register your models here.
from pcari.models import QuantitativeQuestion, QualitativeQuestion, Rating, Comment, UserProgression, GeneralSetting, FlaggedComment, UserData, CommentRating


# admin.site.register(QuantitativeQuestion)
# admin.site.register(QualitativeQuestion)
# admin.site.register(Rating)
# admin.site.register(Comment)
# admin.site.register(UserProgression)
# admin.site.register(UserData)

class Empty:
    def __init__(self):
        self.age = ""
        self.barangay = ""
        self.gender = ""
        self.comment = ""

def dump_comment_ratings_csv(modeladmin, request, queryset):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=malasakit_comment_rating_data.csv'
    writer = csv.writer(response, csv.excel)
    response.write(u'\ufeff'.encode('utf8')) # BOM (optional...Excel needs it to open UTF-8 file properly)
    writer.writerow([
        smart_str(u"User"),
        smart_str(u"User Age"),
        smart_str(u"User Barangay"),
        smart_str(u"User Gender"),
        smart_str(u"Number of ratings"),
        smart_str(u"Average Score"),
        smart_str(u"SE"),
        smart_str(u"Comment"),
        smart_str(u"Date Created"),
    ])

    comment_ratings = CommentRating.objects.all()
    user_data = UserData.objects.all()
    comments = Comment.objects.all()
    for comment in comments:
        try:
            u = user_data.filter(user=comment.user)[0]
        except:
            u = Empty()

        c = comment.comment

        if c == "":
            c = comment.filipino_comment
        # try:
        #     c = comments.filter(id=comment.cid)[0]
        # except:
        #     c = Empty()
        writer.writerow([
                smart_str(comment.user),
                smart_str(u.age),
                smart_str(u.barangay),
                smart_str(u.gender),
                smart_str(comment.number_rated),
                smart_str(comment.average_score),
                smart_str(comment.se),
                smart_str(c),
                smart_str(comment.date),
            ])
    return response
    
dump_comment_ratings_csv.short_description = u"Dump comment ratings as CSV"

def export_comment_csv(modeladmin, request, queryset):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=malasakit_comment_data.csv'
    writer = csv.writer(response, csv.excel)
    response.write(u'\ufeff'.encode('utf8')) # BOM (optional...Excel needs it to open UTF-8 file properly)
    writer.writerow([
        smart_str(u"User"),
        smart_str(u"User Age"),
        smart_str(u"User Barangay"),
        smart_str(u"User Gender"),
        smart_str(u"Comment ID"),
        smart_str(u"Comment"),
        smart_str(u"Filipino Comment"),
        smart_str(u"Average Rcore"),
        smart_str(u"Number Rated"),
        smart_str(u"Date Created"),
        smart_str(u"Original Language"),
    ])
    comments = Comment.objects.all()
    for comment in comments:
        u = user_data.filter(user=comment.user)[0]
        writer.writerow([
            smart_str(comment.user),
            smart_str(u.age),
            smart_str(u.barangay),
            smart_str(u.gender),
            smart_str(comment.id),
            smart_str(comment.comment),
            smart_str(comment.filipino_comment),
            smart_str(comment.average_score),
            smart_str(comment.number_rated),
            smart_str(comment.date),
            smart_str(comment.original_language),
        ])
    return response

export_comment_csv.short_description = u"Export as CSV"

def export_comment_xlsx(modeladmin, request, queryset):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=malasakit_comment_data.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = "Malasakit Data"

    row_num = 0

    columns = [
        (u"User", 15),
        (u"Average_score", 70),
        (u"Number_rated", 70),
    ]

    for col_num in xrange(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
        c.style.font.bold = True
        # set column width
        ws.column_dimensions[get_column_letter(col_num+1)].width = columns[col_num][1]

    for obj in queryset:
        row_num += 1
        row = [
            obj.user,
            obj.average_score,
            obj.number_rated,
        ]
        for col_num in xrange(len(row)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = row[col_num]
            c.style.alignment.wrap_text = True

    wb.save(response)
    return response

def dump_question_ratings_csv(modeladmin, request, queryset):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=malasakit_question_rating_data.csv'
    writer = csv.writer(response, csv.excel)
    response.write(u'\ufeff'.encode('utf8')) # BOM (optional...Excel needs it to open UTF-8 file properly)
    writer.writerow([
        smart_str(u"User"),
        smart_str(u"User Age"),
        smart_str(u"User Barangay"),
        smart_str(u"User Gender"),
        smart_str(u"Question ID"),
        smart_str(u"Score"),
        smart_str(u"Date Created"),
    ])
    ratings = Rating.objects.all()
    user_data = UserData.objects.all()
    # users = User.objects.all()
    for obj in ratings:
        try:
            u = user_data.filter(user=obj.user)[0]
        except:
            u = UserData()

        writer.writerow([
            smart_str(obj.user),
            smart_str(u.age),
            smart_str(u.barangay),
            smart_str(u.gender),
            smart_str(obj.qid),
            smart_str(obj.score),
            smart_str(obj.date),
        ])
    return response

dump_question_ratings_csv.short_description = u"Dump question ratings as CSV"

def export_question_csv(modeladmin, request, queryset):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=malasakit_question_data.csv'
    writer = csv.writer(response, csv.excel)
    response.write(u'\ufeff'.encode('utf8')) # BOM (optional...Excel needs it to open UTF-8 file properly)
    writer.writerow([
        smart_str(u"Question ID"),
        smart_str(u"Question"),
        smart_str(u"Filipino Question"),
        smart_str(u"Average Score"),
        smart_str(u"Number Rated"),
    ])
    questions = QuantitativeQuestion.objects.all()
    for obj in questions:
        writer.writerow([
            smart_str(obj.qid),
            smart_str(obj.question),
            smart_str(obj.filipino_question),
            smart_str(obj.average_score),
            smart_str(obj.number_rated),
        ])
    return response

export_question_csv.short_description = u"Export as CSV"

def export_question_xlsx(modeladmin, request, queryset):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=malasakit_question_data.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = "Malasakit Data"

    row_num = 0

    columns = [
        (u"User", 15),
        (u"Average_score", 70),
        (u"Number_rated", 70),
    ]

    for col_num in xrange(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
        c.style.font.bold = True
        # set column width
        ws.column_dimensions[get_column_letter(col_num+1)].width = columns[col_num][1]

    for obj in queryset:
        row_num += 1
        row = [
            obj.user,
            obj.average_score,
            obj.number_rated,
        ]
        for col_num in xrange(len(row)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = row[col_num]
            c.style.alignment.wrap_text = True

    wb.save(response)
    return response


export_question_xlsx.short_description = u"Export as XLSX"


class GeneralSettingAdmin(admin.ModelAdmin):
    actions = None

admin.site.register(GeneralSetting, GeneralSettingAdmin)


def flag_comment(modeladmin, request, queryset):
	for obj in queryset:
		f = FlaggedComment(user=obj.user,comment=obj.comment,filipino_comment=obj.filipino_comment,date=obj.date,average_score=obj.average_score,number_rated=obj.number_rated,tag=obj.tag)
		f.save()
		obj.delete()

flag_comment.short_description = u"Flag Comment"

def unflag_comment(modeladmin, request, queryset):
	for obj in queryset:
		c = Comment(user=obj.user,comment=obj.comment,filipino_comment=obj.filipino_comment,date=obj.date,average_score=obj.average_score,number_rated=obj.number_rated,tag=obj.tag)
		c.save()
		obj.delete()

unflag_comment.short_description = u"Unflag Comment"


class CommentAdmin(admin.ModelAdmin):
    list_display = ['user', 'comment', 'filipino_comment', 'average_score', 'number_rated', 'se', 'tag', 'original_language']
    list_editable = ['comment','filipino_comment']
    ordering = ['user']
    actions = [flag_comment,dump_comment_ratings_csv]#, export_comment_xlsx]

admin.site.register(Comment, CommentAdmin)

class FlaggedCommentAdmin(admin.ModelAdmin):
    list_display = ['user', 'comment', 'filipino_comment', 'average_score', 'number_rated', 'tag']
    ordering = ['user']
    actions = [unflag_comment,export_comment_csv]#, export_comment_xlsx]

admin.site.register(FlaggedComment, FlaggedCommentAdmin)


class QuantitativeQuestionAdmin(admin.ModelAdmin):
    list_display = ['question', 'filipino_question', 'average_score']
    ordering = ['qid']
    actions = [dump_question_ratings_csv]#,export_question_xlsx]

admin.site.register(QuantitativeQuestion, QuantitativeQuestionAdmin)

# class UserProgressionAdmin(admin.ModelAdmin):
# 	a = 10
# 	list_display = [a]

class UserDataAdmin(admin.ModelAdmin):
    list_display = ['user', 'age', 'barangay', 'gender', 'language']
    ordering = ['user']

admin.site.register(UserData,UserDataAdmin)
# admin.site.register(UserProgression, UserProgressionAdmin)
